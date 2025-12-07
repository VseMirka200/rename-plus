"""Модуль для удаления метаданных из файлов.

Поддерживает удаление метаданных из:
- Изображений (через Pillow)
- Аудио файлов (через mutagen)
- Документов DOCX (через python-docx)
- PDF файлов (через PyPDF2/pypdf)
- XLSX файлов (через openpyxl)
- PPTX файлов (через python-pptx)
"""

import logging
import os
import shutil
from typing import List, Optional, Tuple

logger = logging.getLogger(__name__)


class MetadataRemover:
    """Класс для удаления метаданных из файлов."""
    
    def __init__(self):
        """Инициализация удалителя метаданных."""
        self.pillow_available = False
        self.mutagen_available = False
        self.docx_available = False
        self.pdf_available = False
        self.xlsx_available = False
        self.pptx_available = False
        
        # Попытка импортировать Pillow для работы с изображениями
        try:
            from PIL import Image
            self.Image = Image
            self.pillow_available = True
        except ImportError:
            self.pillow_available = False
        
        # Попытка импортировать mutagen для работы с аудио
        try:
            from mutagen import File as MutagenFile
            self.MutagenFile = MutagenFile
            self.mutagen_available = True
        except ImportError:
            self.mutagen_available = False
        
        # Попытка импортировать python-docx для работы с DOCX
        try:
            from docx import Document
            self.Document = Document
            self.docx_available = True
        except ImportError:
            self.docx_available = False
        
        # Попытка импортировать PyPDF2 для работы с PDF
        try:
            import PyPDF2
            self.PyPDF2 = PyPDF2
            self.pdf_available = True
        except ImportError:
            # Пробуем pypdf (новая версия PyPDF2)
            try:
                import pypdf
                self.PyPDF2 = pypdf
                self.pdf_available = True
            except ImportError:
                self.pdf_available = False
        
        # Попытка импортировать openpyxl для работы с XLSX
        try:
            from openpyxl import load_workbook, Workbook
            self.load_workbook = load_workbook
            self.Workbook = Workbook
            self.xlsx_available = True
        except ImportError:
            self.xlsx_available = False
        
        # Попытка импортировать python-pptx для работы с PPTX
        try:
            from pptx import Presentation
            self.Presentation = Presentation
            self.pptx_available = True
        except ImportError:
            self.pptx_available = False
    
    def can_remove_metadata(self, file_path: str) -> bool:
        """Проверка возможности удаления метаданных из файла.
        
        Args:
            file_path: Путь к файлу
            
        Returns:
            True если можно удалить метаданные, False иначе
        """
        if not os.path.exists(file_path):
            return False
        
        ext = os.path.splitext(file_path)[1].lower()
        
        # Поддерживаемые форматы изображений (SVG исключен, так как Pillow не поддерживает его напрямую)
        image_extensions = {
            '.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.webp', '.gif',
            '.ico', '.jfif', '.jp2', '.jpx', '.j2k', '.j2c', '.pcx', '.ppm', 
            '.pgm', '.pbm', '.pnm', '.psd', '.xbm', '.xpm'
        }
        if ext in image_extensions and self.pillow_available:
            return True
        
        # Поддерживаемые форматы аудио
        audio_extensions = {
            '.mp3', '.flac', '.ogg', '.m4a', '.aac', '.wma', '.wav',
            '.mp4', '.m4p', '.aiff', '.au', '.ra', '.amr', '.3gp', '.opus',
            '.ape', '.mpc', '.tta', '.wv', '.dsf', '.dff', '.mka', '.mkv'
        }
        if ext in audio_extensions and self.mutagen_available:
            return True
        
        # Поддерживаемые форматы документов
        document_extensions = {
            '.docx': self.docx_available,
            '.pdf': self.pdf_available,
            '.xlsx': self.xlsx_available,
            '.pptx': self.pptx_available,
            '.doc': False,  # Старый формат, сложнее обрабатывать
            '.xls': False,  # Старый формат
            '.ppt': False   # Старый формат
        }
        if ext in document_extensions:
            return document_extensions[ext]
        
        return False
    
    def remove_metadata(self, file_path: str, create_backup: bool = True, 
                       remove_options: Optional[dict] = None) -> Tuple[bool, str]:
        """Удаление метаданных из файла.
        
        Args:
            file_path: Путь к файлу
            create_backup: Создавать ли резервную копию перед удалением
            remove_options: Словарь с опциями удаления (author, title, subject, comments, keywords, category, dates, all)
                           Если None, удаляются все метаданные
            
        Returns:
            Кортеж (успех, сообщение)
        """
        # Если опции не указаны, удаляем всё
        if remove_options is None:
            remove_options = {'all': True}
        if not os.path.exists(file_path):
            return False, "Файл не найден"
        
        if not self.can_remove_metadata(file_path):
            return False, "Формат файла не поддерживается для удаления метаданных"
        
        ext = os.path.splitext(file_path)[1].lower()
        
        # Создаем резервную копию если нужно
        backup_path = None
        if create_backup:
            try:
                backup_path = file_path + '.backup'
                shutil.copy2(file_path, backup_path)
            except Exception as e:
                logger.warning(f"Не удалось создать резервную копию: {e}")
        
        try:
            # Удаление метаданных из изображений
            image_extensions = {
                '.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.webp', '.gif',
                '.ico', '.jfif', '.jp2', '.jpx', '.j2k', '.j2c', '.pcx', '.ppm', 
                '.pgm', '.pbm', '.pnm', '.psd', '.xbm', '.xpm'
            }
            if ext in image_extensions and self.pillow_available:
                return self._remove_image_metadata(file_path, backup_path, remove_options)
            
            # Удаление метаданных из аудио
            audio_extensions = {
                '.mp3', '.flac', '.ogg', '.m4a', '.aac', '.wma', '.wav',
                '.mp4', '.m4p', '.aiff', '.au', '.ra', '.amr', '.3gp', '.opus',
                '.ape', '.mpc', '.tta', '.wv', '.dsf', '.dff', '.mka', '.mkv'
            }
            if ext in audio_extensions and self.mutagen_available:
                return self._remove_audio_metadata(file_path, backup_path, remove_options)
            
            # Удаление метаданных из документов
            if ext == '.docx' and self.docx_available:
                return self._remove_docx_metadata(file_path, backup_path, remove_options)
            elif ext == '.pdf' and self.pdf_available:
                return self._remove_pdf_metadata(file_path, backup_path, remove_options)
            elif ext == '.xlsx' and self.xlsx_available:
                return self._remove_xlsx_metadata(file_path, backup_path, remove_options)
            elif ext == '.pptx' and self.pptx_available:
                return self._remove_pptx_metadata(file_path, backup_path, remove_options)
            
            return False, "Формат файла не поддерживается"
            
        except Exception as e:
            logger.error(f"Ошибка при удалении метаданных из {file_path}: {e}", exc_info=True)
            # Восстанавливаем из резервной копии при ошибке
            if backup_path and os.path.exists(backup_path):
                try:
                    shutil.copy2(backup_path, file_path)
                    os.remove(backup_path)
                except Exception:
                    pass
            return False, f"Ошибка: {str(e)}"
    
    def _remove_image_metadata(self, file_path: str, backup_path: Optional[str],
                               remove_options: Optional[dict] = None) -> Tuple[bool, str]:
        """Удаление метаданных из изображения.
        
        Args:
            file_path: Путь к файлу изображения
            backup_path: Путь к резервной копии (если есть)
            remove_options: Опции удаления (для изображений всегда удаляются все EXIF данные)
            
        Returns:
            Кортеж (успех, сообщение)
        """
        # Для изображений всегда удаляются все EXIF данные (нельзя выборочно через Pillow)
        try:
            # Открываем изображение
            with self.Image.open(file_path) as img:
                # Создаем новое изображение без метаданных
                # Конвертируем в RGB для JPEG (если нужно)
                if img.mode in ('RGBA', 'LA', 'P'):
                    # Сохраняем прозрачность для PNG
                    if os.path.splitext(file_path)[1].lower() in ('.png', '.gif', '.webp'):
                        new_img = img.copy()
                    else:
                        # Для JPEG конвертируем в RGB
                        background = self.Image.new('RGB', img.size, (255, 255, 255))
                        if img.mode == 'P':
                            img = img.convert('RGBA')
                        background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                        new_img = background
                else:
                    new_img = img.copy()
                
                # Сохраняем без метаданных
                # Для JPEG используем optimize=True для лучшего сжатия
                save_kwargs = {}
                if os.path.splitext(file_path)[1].lower() in ('.jpg', '.jpeg'):
                    save_kwargs['quality'] = 95
                    save_kwargs['optimize'] = True
                
                new_img.save(file_path, **save_kwargs)
            
            # Удаляем резервную копию если операция успешна
            if backup_path and os.path.exists(backup_path):
                try:
                    os.remove(backup_path)
                except Exception:
                    pass
            
            return True, "Метаданные успешно удалены"
            
        except Exception as e:
            logger.error(f"Ошибка при удалении метаданных изображения: {e}", exc_info=True)
            return False, f"Ошибка: {str(e)}"
    
    def _remove_audio_metadata(self, file_path: str, backup_path: Optional[str],
                               remove_options: Optional[dict] = None) -> Tuple[bool, str]:
        """Удаление метаданных из аудио файла.
        
        Args:
            file_path: Путь к аудио файлу
            backup_path: Путь к резервной копии (если есть)
            remove_options: Опции удаления (для аудио всегда удаляются все теги)
            
        Returns:
            Кортеж (успех, сообщение)
        """
        try:
            audio_file = self.MutagenFile(file_path)
            if audio_file is None:
                return False, "Не удалось открыть аудио файл"
            
            # Для аудио всегда удаляются все теги (нельзя выборочно через mutagen)
            audio_file.delete()
            audio_file.save()
            
            # Удаляем резервную копию если операция успешна
            if backup_path and os.path.exists(backup_path):
                try:
                    os.remove(backup_path)
                except Exception:
                    pass
            
            return True, "Метаданные успешно удалены"
            
        except Exception as e:
            logger.error(f"Ошибка при удалении метаданных аудио: {e}", exc_info=True)
            return False, f"Ошибка: {str(e)}"
    
    def _remove_docx_metadata(self, file_path: str, backup_path: Optional[str],
                              remove_options: Optional[dict] = None) -> Tuple[bool, str]:
        """Удаление метаданных из DOCX файла.
        
        Args:
            file_path: Путь к DOCX файлу
            backup_path: Путь к резервной копии (если есть)
            remove_options: Опции удаления (author, title, subject, comments, keywords, category, dates, all)
            
        Returns:
            Кортеж (успех, сообщение)
        """
        if remove_options is None:
            remove_options = {'all': True}
        
        try:
            # Открываем документ
            doc = self.Document(file_path)
            
            # Очищаем основные свойства документа
            core_props = doc.core_properties
            
            # Удаляем свойства в зависимости от опций
            if remove_options.get('all', False) or remove_options.get('author', False):
                try:
                    core_props.author = ''
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('title', False):
                try:
                    core_props.title = ''
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('subject', False):
                try:
                    core_props.subject = ''
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('comments', False):
                try:
                    core_props.comments = ''
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('category', False):
                try:
                    core_props.category = ''
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('keywords', False):
                try:
                    core_props.keywords = ''
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('description', False):
                try:
                    # В DOCX нет отдельного поля description, используем comments
                    if not remove_options.get('comments', False):
                        core_props.comments = ''
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('last_modified', False):
                try:
                    core_props.last_modified_by = ''
                except (AttributeError, TypeError):
                    pass
            
            # revision должен быть положительным целым числом, устанавливаем в 1
            if remove_options.get('all', False) or remove_options.get('revision', False):
                try:
                    core_props.revision = 1
                except (AttributeError, TypeError):
                    pass
            
            # Даты не устанавливаем в None, так как они требуют datetime объект
            # Если нужно удалить даты, можно установить текущую дату, но лучше не трогать
            # Для dates опции можно было бы установить текущую дату, но это может быть нежелательно
            
            # Сохраняем документ
            doc.save(file_path)
            
            # Удаляем резервную копию если операция успешна
            if backup_path and os.path.exists(backup_path):
                try:
                    os.remove(backup_path)
                except Exception:
                    pass
            
            return True, "Метаданные успешно удалены"
            
        except Exception as e:
            logger.error(f"Ошибка при удалении метаданных DOCX: {e}", exc_info=True)
            return False, f"Ошибка: {str(e)}"
    
    def _remove_pdf_metadata(self, file_path: str, backup_path: Optional[str],
                             remove_options: Optional[dict] = None) -> Tuple[bool, str]:
        """Удаление метаданных из PDF файла.
        
        Args:
            file_path: Путь к PDF файлу
            backup_path: Путь к резервной копии (если есть)
            remove_options: Опции удаления (для PDF всегда удаляются все метаданные)
            
        Returns:
            Кортеж (успех, сообщение)
        """
        # Для PDF всегда удаляются все метаданные (PyPDF2 не поддерживает выборочное удаление)
        try:
            # Открываем PDF
            with open(file_path, 'rb') as input_file:
                pdf_reader = self.PyPDF2.PdfReader(input_file)
                pdf_writer = self.PyPDF2.PdfWriter()
                
                # Копируем все страницы
                for page in pdf_reader.pages:
                    pdf_writer.add_page(page)
                
                # Удаляем метаданные
                pdf_writer.add_metadata({})
            
            # Сохраняем новый PDF
            with open(file_path, 'wb') as output_file:
                pdf_writer.write(output_file)
            
            # Удаляем резервную копию если операция успешна
            if backup_path and os.path.exists(backup_path):
                try:
                    os.remove(backup_path)
                except Exception:
                    pass
            
            return True, "Метаданные успешно удалены"
            
        except Exception as e:
            logger.error(f"Ошибка при удалении метаданных PDF: {e}", exc_info=True)
            return False, f"Ошибка: {str(e)}"
    
    def _remove_xlsx_metadata(self, file_path: str, backup_path: Optional[str],
                              remove_options: Optional[dict] = None) -> Tuple[bool, str]:
        """Удаление метаданных из XLSX файла.
        
        Args:
            file_path: Путь к XLSX файлу
            backup_path: Путь к резервной копии (если есть)
            remove_options: Опции удаления (author, title, subject, comments, keywords, category, dates, all)
            
        Returns:
            Кортеж (успех, сообщение)
        """
        if remove_options is None:
            remove_options = {'all': True}
        
        try:
            # Открываем книгу
            wb = self.load_workbook(file_path)
            
            # Очищаем свойства документа в зависимости от опций
            if remove_options.get('all', False) or remove_options.get('author', False):
                try:
                    wb.properties.creator = None
                    wb.properties.lastModifiedBy = None
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('title', False):
                try:
                    wb.properties.title = None
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('subject', False):
                try:
                    wb.properties.subject = None
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('description', False):
                try:
                    wb.properties.description = None
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('comments', False):
                try:
                    wb.properties.comments = None
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('keywords', False):
                try:
                    wb.properties.keywords = None
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('category', False):
                try:
                    wb.properties.category = None
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('revision', False):
                try:
                    # В XLSX нет прямого поля revision, но можно очистить другие свойства
                    pass
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('last_modified', False):
                try:
                    # lastModifiedBy уже обрабатывается выше
                    pass
                except (AttributeError, TypeError):
                    pass
            
            # Сохраняем книгу
            wb.save(file_path)
            
            # Удаляем резервную копию если операция успешна
            if backup_path and os.path.exists(backup_path):
                try:
                    os.remove(backup_path)
                except Exception:
                    pass
            
            return True, "Метаданные успешно удалены"
            
        except Exception as e:
            logger.error(f"Ошибка при удалении метаданных XLSX: {e}", exc_info=True)
            return False, f"Ошибка: {str(e)}"
    
    def _remove_pptx_metadata(self, file_path: str, backup_path: Optional[str],
                              remove_options: Optional[dict] = None) -> Tuple[bool, str]:
        """Удаление метаданных из PPTX файла.
        
        Args:
            file_path: Путь к PPTX файлу
            backup_path: Путь к резервной копии (если есть)
            remove_options: Опции удаления (author, title, subject, comments, keywords, category, dates, all)
            
        Returns:
            Кортеж (успех, сообщение)
        """
        if remove_options is None:
            remove_options = {'all': True}
        
        try:
            # Открываем презентацию
            prs = self.Presentation(file_path)
            
            # Очищаем основные свойства в зависимости от опций
            core_props = prs.core_properties
            
            if remove_options.get('all', False) or remove_options.get('author', False):
                try:
                    core_props.author = None
                    core_props.last_modified_by = None
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('title', False):
                try:
                    core_props.title = None
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('subject', False):
                try:
                    core_props.subject = None
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('comments', False):
                try:
                    core_props.comments = None
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('category', False):
                try:
                    core_props.category = None
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('keywords', False):
                try:
                    core_props.keywords = None
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('description', False):
                try:
                    # В PPTX нет отдельного поля description, используем comments
                    if not remove_options.get('comments', False):
                        core_props.comments = None
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('revision', False):
                try:
                    # В PPTX нет прямого поля revision
                    pass
                except (AttributeError, TypeError):
                    pass
            
            if remove_options.get('all', False) or remove_options.get('last_modified', False):
                try:
                    # last_modified_by уже обрабатывается выше
                    pass
                except (AttributeError, TypeError):
                    pass
            
            # Сохраняем презентацию
            prs.save(file_path)
            
            # Удаляем резервную копию если операция успешна
            if backup_path and os.path.exists(backup_path):
                try:
                    os.remove(backup_path)
                except Exception:
                    pass
            
            return True, "Метаданные успешно удалены"
            
        except Exception as e:
            logger.error(f"Ошибка при удалении метаданных PPTX: {e}", exc_info=True)
            return False, f"Ошибка: {str(e)}"
    
    def remove_metadata_batch(self, file_paths: List[str], create_backup: bool = True) -> List[Tuple[str, bool, str]]:
        """Удаление метаданных из нескольких файлов.
        
        Args:
            file_paths: Список путей к файлам
            create_backup: Создавать ли резервные копии
            
        Returns:
            Список кортежей (путь, успех, сообщение)
        """
        results = []
        for file_path in file_paths:
            success, message = self.remove_metadata(file_path, create_backup)
            results.append((file_path, success, message))
        return results

